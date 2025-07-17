import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

class Report:

    def __init__(self, kontakte):
        # Tkinter root erstellen und verbergen
        root = tk.Tk()
        root.withdraw()

        # Dateidialog öffnen, um Speicherort und Dateinamen auszuwählen
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Speichern unter"
        )

        # Wenn der Benutzer den Dialog abbricht, wird keine Datei erstellt
        if not file_path:
            return

        # Excel Workbook erstellen
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Visa Report"

        headers = [
            "إسم الموظف",  # Employee
            "رقم التسجيل",  # Visa Number
            "الأسم بالأنجليزي",  # Vorname
            "الأسم بالعربي",  # Nachname
            "الرقم الوطني",  # Passport
            "المهنة",  # Profession
            "تاريخ الإصدار",  # Visa Issued
           "اسم الأم",  # Duration
            "تاريخ الميلاد",  # Visa Valid
            "أسباب التواجد",  # Entries Number
            "مكان أصدار الوثيقة",  # Entry Fee
            "محل الميلاد",  # Visa Art
            "العنوان بليبيا",  # Port
            "أقرب الأقارب بليبيا",  # Work
            "أسباب صرف الوثيقة"  # Note
        ]

        # Hinzufügen der Tabellenkopfzeile
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

        # Hinzufügen der Kontakte zur Tabelle
        #todo für reihnfolge
        for row_num, kontakt in enumerate(kontakte, 2):
            sheet.cell(row=row_num, column=1, value=kontakt.employee)
            sheet.cell(row=row_num, column=2, value=kontakt.visaNumber)
            sheet.cell(row=row_num, column=3, value=kontakt.vorname)
            sheet.cell(row=row_num, column=4, value=kontakt.nachname)
            sheet.cell(row=row_num, column=5, value=kontakt.passport)
            sheet.cell(row=row_num, column=6, value=kontakt.profession)
            sheet.cell(row=row_num, column=7, value=kontakt.visaIss)
            sheet.cell(row=row_num, column=8, value=kontakt.duration)
            sheet.cell(row=row_num, column=9, value=kontakt.visaValid)
            sheet.cell(row=row_num, column=10, value=kontakt.entriesNumber)
            sheet.cell(row=row_num, column=11, value=kontakt.entryFee)
            sheet.cell(row=row_num, column=12, value=kontakt.visaArt)
            sheet.cell(row=row_num, column=13, value=kontakt.port)
            sheet.cell(row=row_num, column=14, value=kontakt.work)
            sheet.cell(row=row_num, column=15, value=kontakt.note)

            # Zellen zentrieren
            for col_num in range(1, 8):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

        # Spaltenbreite anpassen
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Excel-Datei speichern
        workbook.save(file_path)

